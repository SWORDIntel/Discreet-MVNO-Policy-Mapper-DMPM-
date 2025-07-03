"""
Export functionality for GHOST DMPM
Formats: CSV, JSON, Excel, PDF, HTML
"""

import csv
import json
import logging # For logging within the class
from pathlib import Path
from typing import List, Dict, Any, Optional

# Placeholder for GhostConfig, assuming it's available
# from ghost_dmpm.core.config import GhostConfig

# Third-party libraries to be added to requirements if used:
# import openpyxl # For Excel
# from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle # For PDF
# from reportlab.lib.styles import getSampleStyleSheet # For PDF
# from reportlab.lib import colors # For PDF
# from reportlab.lib.units import inch # For PDF
# from jinja2 import Environment, FileSystemLoader # For HTML

class GhostExporter:
    """
    Handles exporting data to various formats like CSV, JSON, Excel, PDF, HTML.
    """
    def __init__(self, config: Any): # Using Any for config type for now
        """
        Initializes the GhostExporter system.

        Args:
            config (GhostConfig): The system's configuration object.
                                  (Using Any for now to avoid direct dependency if not strictly needed by all methods)
        """
        self.config = config
        # It's good practice to have a logger, even if config is Any for now.
        # If config is a real GhostConfig, this will work. Otherwise, it might need adjustment.
        if hasattr(config, 'get_logger'):
            self.logger = config.get_logger("GhostExporter")
        else:
            self.logger = logging.getLogger("GhostExporter")
            if not self.logger.handlers: # Basic config if no handlers exist
                 logging.basicConfig(level=logging.INFO)


    def export_json(self, data: Dict[str, Any], output_path: Path, pretty: bool = True, compress: bool = False) -> Optional[Path]:
        """
        Exports data to a JSON file.

        Args:
            data (Dict[str, Any]): The data to export.
            output_path (Path): The path to save the JSON file (e.g., Path("report.json")).
            pretty (bool): If True, formats the JSON output with indentation. Defaults to True.
            compress (bool): If True, and if the filename ends with .gz, it will compress the JSON. (Not yet implemented)

        Returns:
            Optional[Path]: The path to the saved file, or None if an error occurred.
        """
        # TODO: Implement compression if compress=True and output_path.suffix == '.gz'
        if compress:
            self.logger.warning("JSON compression is not yet implemented in export_json.")

        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, 'w', encoding='utf-8') as f:
                if pretty:
                    json.dump(data, f, indent=4, ensure_ascii=False)
                else:
                    json.dump(data, f, ensure_ascii=False)
            self.logger.info(f"Data successfully exported to JSON: {output_path}")
            return output_path
        except IOError as e:
            self.logger.error(f"IOError exporting data to JSON {output_path}: {e}", exc_info=True)
        except Exception as e:
            self.logger.error(f"Unexpected error exporting data to JSON {output_path}: {e}", exc_info=True)
        return None

    def export_csv(self, data: List[Dict[str, Any]], output_path: Path, columns: Optional[List[str]] = None) -> Optional[Path]:
        """
        Exports data (list of dictionaries) to a CSV file.

        Args:
            data (List[Dict[str, Any]]): A list of dictionaries, where each dictionary represents a row.
            output_path (Path): The path to save the CSV file (e.g., Path("report.csv")).
            columns (Optional[List[str]]): A list of column names (keys from the dictionaries) to include
                                           and their order. If None, all keys from the first item are used.

        Returns:
            Optional[Path]: The path to the saved file, or None if an error occurred.
        """
        if not data:
            self.logger.warning("No data provided for CSV export.")
            return None

        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)

            if columns is None:
                columns = list(data[0].keys()) if data else []

            if not columns:
                 self.logger.warning("No columns determined for CSV export.")
                 return None

            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=columns, extrasaction='ignore')
                writer.writeheader()
                for row in data:
                    writer.writerow(row)

            self.logger.info(f"Data successfully exported to CSV: {output_path}")
            return output_path
        except IOError as e:
            self.logger.error(f"IOError exporting data to CSV {output_path}: {e}", exc_info=True)
        except Exception as e:
            self.logger.error(f"Unexpected error exporting data to CSV {output_path}: {e}", exc_info=True)
        return None

    def export_excel(self, data: Dict[str, List[Dict[str, Any]]], output_path: Path) -> Optional[Path]:
        """
        Exports data to a multi-sheet Excel workbook.
        Each key in the `data` dictionary will be a sheet name, and the value (list of dicts) its content.

        Args:
            data (Dict[str, List[Dict[str, Any]]]): A dictionary where keys are sheet names
                                                     and values are lists of dictionaries for rows.
            output_path (Path): The path to save the Excel file (e.g., Path("report.xlsx")).

        Returns:
            Optional[Path]: The path to the saved file, or None if an error occurred.
        """
        self.logger.warning("Excel export is not yet implemented.")
        # Example structure:
        # try:
        #     import openpyxl
        #     wb = openpyxl.Workbook()
        #     default_sheet = wb.active # Get default sheet
        #     wb.remove(default_sheet) # Remove it if we are creating named sheets
        #
        #     for sheet_name, sheet_data in data.items():
        #         ws = wb.create_sheet(title=sheet_name)
        #         if not sheet_data:
        #             continue
        #         headers = list(sheet_data[0].keys())
        #         ws.append(headers)
        #         for row_dict in sheet_data:
        #             row_values = [row_dict.get(h) for h in headers]
        #             ws.append(row_values)
        #
        #     output_path.parent.mkdir(parents=True, exist_ok=True)
        #     wb.save(output_path)
        #     self.logger.info(f"Data successfully exported to Excel: {output_path}")
        #     return output_path
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            # Remove default sheet if creating named sheets, or rename it
            if wb.sheetnames == ['Sheet'] and data: # if default sheet exists and we have data for new sheets
                default_sheet = wb.active
                if default_sheet: # Should exist
                    wb.remove(default_sheet)

            first_sheet = True
            for sheet_name, sheet_data_list in data.items():
                ws = wb.create_sheet(title=sheet_name)

                if not sheet_data_list:
                    self.logger.warning(f"No data provided for Excel sheet: {sheet_name}")
                    continue # Create empty sheet or skip? For now, create empty.

                if not isinstance(sheet_data_list, list) or not all(isinstance(item, dict) for item in sheet_data_list):
                    self.logger.error(f"Data for sheet '{sheet_name}' is not a list of dictionaries.")
                    # Optionally, create a sheet with an error message
                    ws.cell(row=1, column=1, value=f"Error: Data for sheet '{sheet_name}' is malformed.")
                    continue

                if not sheet_data_list: # Double check after type check
                    continue

                headers = list(sheet_data_list[0].keys())
                ws.append(headers) # Add header row

                for row_dict in sheet_data_list:
                    row_values = []
                    for h in headers:
                        val = row_dict.get(h)
                        # openpyxl cannot handle complex types like dicts/lists in cells directly
                        if isinstance(val, (dict, list)):
                            try:
                                val = json.dumps(val) # Convert to JSON string
                            except TypeError:
                                val = str(val) # Fallback to generic string
                        row_values.append(val)
                    ws.append(row_values)

                # Auto-adjust column widths (optional, can be slow for large sheets)
                for col_idx, column_cells in enumerate(ws.columns):
                    max_length = 0
                    column = get_column_letter(col_idx + 1)
                    for cell in column_cells:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass # Ignore errors for non-stringifiable values
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = min(adjusted_width, 50) # Cap width

            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            self.logger.info(f"Data successfully exported to Excel: {output_path}")
            return output_path
        except ImportError:
            self.logger.error("openpyxl library is required for Excel export. Please install it (e.g., pip install openpyxl).")
        except Exception as e:
            self.logger.error(f"Error exporting data to Excel {output_path}: {e}", exc_info=True)
        return None

    def export_pdf(self, report_content: Dict[str, Any], output_path: Path) -> Optional[Path]:
        """
        Exports a structured report to a PDF file.

        Args:
            report_content (Dict[str, Any]): A dictionary containing structured report data
                                             (e.g., title, sections, tables, text).
            output_path (Path): The path to save the PDF file (e.g., Path("report.pdf")).

        Returns:
            Optional[Path]: The path to the saved file, or None if an error occurred.
        """
        self.logger.warning("PDF export is not yet implemented.")
        # Example structure using reportlab:
        # try:
        #     from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        #     from reportlab.lib.styles import getSampleStyleSheet
        #     from reportlab.lib import colors
        #     from reportlab.lib.units import inch
        #
        #     output_path.parent.mkdir(parents=True, exist_ok=True)
        #     doc = SimpleDocTemplate(str(output_path))
        #     styles = getSampleStyleSheet()
        #     story = []
        #
        #     story.append(Paragraph(report_content.get("title", "GHOST DMPM Report"), styles['h1']))
        #     story.append(Spacer(1, 0.2*inch))
        #
        #     # Add more content based on report_content structure
        #     # e.g., sections with text, tables
        #
        #     doc.build(story)
        #     self.logger.info(f"Report successfully exported to PDF: {output_path}")
        #     return output_path
        try:
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.lib.enums import TA_LEFT, TA_CENTER

            output_path.parent.mkdir(parents=True, exist_ok=True)
            doc = SimpleDocTemplate(str(output_path))
            styles = getSampleStyleSheet()

            # Add a default style for normal text if not present or customize
            if 'Normal_Left' not in styles:
                styles.add(ParagraphStyle(name='Normal_Left', parent=styles['Normal'], alignment=TA_LEFT))

            story = []

            title_text = report_content.get("title", "GHOST DMPM Report")
            story.append(Paragraph(title_text, styles['h1']))
            story.append(Spacer(1, 0.2 * inch))

            for key, value in report_content.items():
                if key == "title":
                    continue

                story.append(Paragraph(key.replace('_', ' ').title(), styles['h2']))
                story.append(Spacer(1, 0.1 * inch))

                if isinstance(value, str):
                    story.append(Paragraph(value, styles['Normal_Left']))
                elif isinstance(value, list) and value and all(isinstance(i, dict) for i in value):
                    # Create a table for list of dicts
                    table_data = []
                    headers = list(value[0].keys())
                    table_data.append([Paragraph(str(h), styles['Normal']) for h in headers]) # Header row

                    for item_dict in value:
                        row = [Paragraph(str(item_dict.get(h, '')), styles['Normal_Left']) for h in headers]
                        table_data.append(row)

                    if table_data:
                        col_widths = [doc.width/len(headers)] * len(headers) # Distribute width
                        table = Table(table_data, colWidths=col_widths)
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        story.append(table)
                elif isinstance(value, dict):
                    # For dicts, maybe a simple key-value list or pretty print JSON-like
                    for sub_key, sub_value in value.items():
                        story.append(Paragraph(f"<b>{sub_key.replace('_', ' ').title()}:</b> {sub_value}", styles['Normal_Left']))
                        story.append(Spacer(1, 0.05 * inch))
                else: # Fallback for other data types
                    story.append(Paragraph(str(value), styles['Normal_Left']))

                story.append(Spacer(1, 0.2 * inch))

            doc.build(story)
            self.logger.info(f"Report successfully exported to PDF: {output_path}")
            return output_path
        except ImportError:
            self.logger.error("reportlab library is required for PDF export. Please install it (e.g., pip install reportlab).")
        except Exception as e:
            self.logger.error(f"Error exporting report to PDF {output_path}: {e}", exc_info=True)
        return None

    def export_html(self, report_content: Dict[str, Any], output_path: Path, template_name: Optional[str] = None) -> Optional[Path]:
        """
        Exports a structured report to an HTML file using a template.

        Args:
            report_content (Dict[str, Any]): Data to be passed to the HTML template.
            output_path (Path): The path to save the HTML file (e.g., Path("report.html")).
            template_name (Optional[str]): The name of the Jinja2 template to use.
                                           If None, a default basic HTML structure will be attempted.

        Returns:
            Optional[Path]: The path to the saved file, or None if an error occurred.
        """
        self.logger.warning("HTML export is not yet implemented.")
        # Example structure using Jinja2:
        # try:
        #     from jinja2 import Environment, FileSystemLoader #, select_autoescape
        #
        #     # Assuming templates are in a 'templates/export/' directory relative to this file or project root
        #     # This needs to be configured properly.
        #     # template_dir = Path(__file__).parent / 'templates' # Or from config
        #     # env = Environment(loader=FileSystemLoader(template_dir), autoescape=select_autoescape(['html', 'xml']))
        #
        #     # if template_name is None:
        #     #     # Create a very basic HTML string if no template
        #     #     html_output = f"<html><head><title>{report_content.get('title', 'Report')}</title></head>"
        #     #     html_output += f"<body><h1>{report_content.get('title', 'Report')}</h1>"
        #     #     html_output += f"<pre>{json.dumps(report_content, indent=2)}</pre></body></html>"
        #     # else:
        #     #     template = env.get_template(template_name)
        #     #     html_output = template.render(report_content)
        #
        #     # output_path.parent.mkdir(parents=True, exist_ok=True)
        #     # with open(output_path, 'w', encoding='utf-8') as f:
        #     #     f.write(html_output)
        #     # self.logger.info(f"Report successfully exported to HTML: {output_path}")
        #     # return output_path
        try:
            from jinja2 import Environment, FileSystemLoader, select_autoescape

            # Determine template directory - this needs a robust way to find a 'templates' folder.
            # For now, let's assume a 'templates/export/' subdir relative to this file,
            # or a path specified in config.
            template_dir = None
            if hasattr(self.config, 'get_template_dir'): # Ideal: config provides template path
                template_dir = self.config.get_template_dir('export')

            if not template_dir: # Fallback if not in config
                # A common pattern is to have templates near the module or in a project-level dir
                base_template_path = Path(__file__).parent / "export_templates" # Example
                if base_template_path.is_dir():
                     template_dir = base_template_path
                else: # Final fallback or if project structure dictates another location
                    if hasattr(self.config, 'project_root'):
                        template_dir = self.config.project_root / "templates" / "export"
                    else: # Absolute last resort if no project_root
                        template_dir = Path("templates") / "export"


            html_output = ""
            if template_name and template_dir:
                if not Path(template_dir).is_dir():
                    self.logger.warning(f"Template directory '{template_dir}' not found for HTML export. Falling back to basic HTML.")
                    template_name = None # Force fallback
                else:
                    try:
                        env = Environment(
                            loader=FileSystemLoader(str(template_dir)),
                            autoescape=select_autoescape(['html', 'xml'])
                        )
                        template = env.get_template(template_name)
                        html_output = template.render(report_content)
                    except Exception as e:
                        self.logger.error(f"Error loading/rendering Jinja2 template '{template_name}': {e}. Falling back to basic HTML.")
                        template_name = None # Force fallback to basic if template fails

            if not template_name:  # If no template_name or if template loading failed
                # Create a very basic HTML string if no template
                title = report_content.get("title", "GHOST DMPM Report")
                html_output = f"<html><head><title>{title}</title>"
                html_output += "<style>body{font-family: sans-serif;} table{border-collapse: collapse; width: 100%;} th, td{border: 1px solid #ddd; padding: 8px; text-align: left;} th{background-color: #f2f2f2;}</style>"
                html_output += f"</head><body><h1>{title}</h1>"

                # Basic rendering for common data structures
                for key, value in report_content.items():
                    if key == "title": continue
                    html_output += f"<h2>{key.replace('_', ' ').title()}</h2>"
                    if isinstance(value, list) and value and all(isinstance(i, dict) for i in value):
                        html_output += "<table><thead><tr>"
                        for header in value[0].keys():
                            html_output += f"<th>{header}</th>"
                        html_output += "</tr></thead><tbody>"
                        for item_dict in value:
                            html_output += "<tr>"
                            for K_item, v_item in item_dict.items():
                                html_output += f"<td>{v_item}</td>"
                            html_output += "</tr>"
                        html_output += "</tbody></table>"
                    elif isinstance(value, dict):
                        html_output += f"<pre>{json.dumps(value, indent=2)}</pre>"
                    else:
                         html_output += f"<p>{value}</p>"
                html_output += "</body></html>"

            output_path.parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_output)
            self.logger.info(f"Report successfully exported to HTML: {output_path}")
            return output_path
        except ImportError:
            self.logger.error("Jinja2 library is required for HTML export. Please install it (e.g., pip install Jinja2).")
        except Exception as e:
            self.logger.error(f"Error exporting report to HTML {output_path}: {e}", exc_info=True)
        return None

# Example usage:
if __name__ == '__main__':
    # Mock GhostConfig for example
    class MockConfig:
        def get_logger(self, name):
            logger = logging.getLogger(name)
            if not logger.handlers:
                logging.basicConfig(level=logging.INFO)
            return logger

    mock_config = MockConfig()
    exporter = GhostExporter(config=mock_config)

    # Create dummy data
    test_data_list = [
        {'id': 1, 'name': 'Alice', 'value': 100, 'city': 'New York'},
        {'id': 2, 'name': 'Bob', 'value': 200, 'notes': 'Test note'},
        {'id': 3, 'name': 'Charlie', 'value': 150, 'city': 'London', 'active': True}
    ]
    test_data_dict = {"summary": {"count": 3, "total_value": 450}, "items": test_data_list}

    export_dir = Path("test_exports")
    export_dir.mkdir(exist_ok=True)

    # Test JSON export
    json_path = exporter.export_json(test_data_dict, export_dir / "test_report.json")
    if json_path:
        print(f"JSON export successful: {json_path}")
        json_path_ugly = exporter.export_json(test_data_dict, export_dir / "test_report.ugly.json", pretty=False)
        print(f"Ugly JSON export successful: {json_path_ugly}")

    # Test CSV export
    csv_path = exporter.export_csv(test_data_list, export_dir / "test_report.csv")
    if csv_path:
        print(f"CSV export successful: {csv_path}")

    csv_path_cols = exporter.export_csv(
        test_data_list,
        export_dir / "test_report_cols.csv",
        columns=['name', 'city', 'value'] # Specific columns
    )
    if csv_path_cols:
        print(f"CSV export with specific columns successful: {csv_path_cols}")

    # Placeholders for other exports
    print("\nExcel, PDF, HTML export methods are placeholders.")
    # excel_data = {"Sheet1": test_data_list, "Summary": [{"total_items": len(test_data_list)}]}
    # exporter.export_excel(excel_data, export_dir / "test_report.xlsx")
    #
    # pdf_report_data = {"title": "My PDF Report", "content": "This is the content for the PDF."}
    # exporter.export_pdf(pdf_report_data, export_dir / "test_report.pdf")
    #
    # html_report_data = {"title": "My HTML Report", "items": test_data_list}
    # exporter.export_html(html_report_data, export_dir / "test_report.html")

    print(f"\nCheck the '{export_dir}' directory for output files.")

# Add to src/ghost_dmpm/enhancements/__init__.py:
# from .export import GhostExporter
# __all__ = [..., 'GhostExporter']
